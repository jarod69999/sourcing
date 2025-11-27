import streamlit as st
import pandas as pd
import re, time, unicodedata
from io import BytesIO
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
from openpyxl import load_workbook
import folium
from folium.features import DivIcon
from streamlit.components.v1 import html as st_html

# ========================== CONFIG ==========================
TEMPLATE_PATH = "Sourcing base.xlsx"   # modèle Excel avec en-têtes
START_ROW = 11                         # 1re ligne de data dans le modèle

PRIMARY = "#0b1d4f"
BG      = "#f5f0eb"
st.set_page_config(page_title="MOA – v2 ", page_icon="📍", layout="wide")
# ===============================================================
# KEEP ALIVE – empêche l'app de se mettre en sommeil (ping interne)
# ===============================================================
keepalive_js = """
<script>
    function keepAlive() {
        fetch("/_stcore/health", {method:"GET"});
    }
    setInterval(keepAlive, 300000);  // 300 000 ms = 5 minutes
</script>
"""
import streamlit as st
st.markdown(keepalive_js, unsafe_allow_html=True)

st.markdown(f"""
<style>
 .stApp {{background:{BG};font-family:Inter,system-ui,Roboto,Arial;}}
 h1,h2,h3{{color:{PRIMARY};}}
 .stDownloadButton > button{{background:{PRIMARY};color:#fff;border-radius:8px;border:0;}}
 .stTextInput > div > div > input{{background:#fff;}}
 .stFileUploader label div{{background:#fff;}}
</style>
""", unsafe_allow_html=True)

# ====================== GEO & HELPERS =======================
COUNTRY_WORDS = {
    "france","belgique","belgium","belgie","belgië","espagne","españa","portugal",
    "italie","italia","deutschland","germany","suisse","switzerland","luxembourg",
    "pays-bas","pays bas","netherlands","nederland"
}
CP_FALLBACK_RE = re.compile(r"\b\d{4,6}\b")
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

INDUS_TOKENS = ["implant-indus-2","implant-indus-3","implant-indus-4","implant-indus-5"]
HQ_TOKEN     = "adresse-du-siège"

import requests

# ================== VERIFICATION CLE ORS ==================

def ors_distance(coord1, coord2, ors_key=""):
    """
    Essaie de calculer la distance routière (driving-car) via OpenRouteService.
    Si la requête échoue ou que la clé est absente, renvoie None.
    """
    if not coord1 or not coord2 or not ors_key:
        return None
    url = "https://api.openrouteservice.org/v2/directions/driving-car"
    headers = {"Authorization": ors_key, "Content-Type": "application/json"}
    data = {"coordinates": [[coord1[1], coord1[0]], [coord2[1], coord2[0]]]}
    try:
        r = requests.post(url, json=data, headers=headers, timeout=30)
        if r.status_code == 200:
            js = r.json()
            return js["routes"][0]["summary"]["distance"] / 1000.0  # km
        else:
            print(f"⚠️ ORS error {r.status_code}: {r.text[:200]}")
    except Exception as e:
        print(f"⚠️ ORS request failed: {e}")
    return None


def _norm(text: str) -> str:
    if not isinstance(text,str): return ""
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("’","'").replace("–","-").replace("—","-")
    text = re.sub(r"\s+", " ", text).strip()
    return text

def _fix_postcode_spaces(text: str) -> str:
    # "40 300" -> "40300", "75 018" -> "75018"
    return re.sub(r"\b(\d{2})\s?(\d{3})\b", r"\1\2", text)

def has_explicit_country(s: str) -> bool:
    return any(w in s.lower() for w in COUNTRY_WORDS)

def extract_cp_fallback(text: str) -> str:
    if not isinstance(text, str): return ""
    t = _fix_postcode_spaces(_norm(text))
    m = CP_FALLBACK_RE.search(t)
    return m.group(0) if m else ""

def extract_cp_city(text: str):
    """Essaie d'extraire (cp, ville) FR/BE à partir de l'adresse brute."""
    if not isinstance(text,str): return ("","")
    t = _fix_postcode_spaces(_norm(text))
    # pattern 1: '40300 Hastingues'
    m = re.search(r"\b(\d{4,5})\b[ ,\-]*([A-Za-zÀ-ÖØ-öø-ÿ' \-]{2,})", t)
    if m:
        cp = m.group(1)
        ville = m.group(2).split(",")[0].strip()
        ville = re.sub(r"\bcedex\b.*$", "", ville, flags=re.I).strip()
        return (cp, ville)
    # pattern 2: 'Hastingues 40300'
    m = re.search(r"([A-Za-zÀ-ÖØ-öø-ÿ' \-]{2,})[ ,\-]*(\d{4,5})\b", t)
    if m:
        ville = m.group(1).split(",")[0].strip()
        ville = re.sub(r"\bcedex\b.*$", "", ville, flags=re.I).strip()
        return (m.group(2), ville)
    return ("","")

def clean_street_numbers(addr: str) -> str:
    """
    Si un numéro à 3-4 chiffres est au début et qu'un code postal FR à 5 chiffres apparaît plus loin,
    on supprime le premier pour éviter la confusion (ex: '1070 Route de...' => 'Route de...').
    """
    if not isinstance(addr, str):
        return addr
    addr = addr.strip()
    # Si code postal à 5 chiffres quelque part, supprimer le nombre initial à 3–4 chiffres
    if re.search(r"\b\d{5}\b", addr):
        addr = re.sub(r"^\s*\d{3,4}\b\s*", "", addr)
    return addr


def clean_internal_codes(addr: str) -> str:
    """Nettoie BP, CS et espaces inutiles."""
    if not isinstance(addr, str):
        return addr
    addr = re.sub(r"\b(CS|BP)\s*\d{3,6}\b", "", addr, flags=re.IGNORECASE)
    addr = re.sub(r"[-]{2,}", "-", addr)
    addr = re.sub(r"\s{2,}", " ", addr).strip(" ,.-")
    return addr


@st.cache_data(show_spinner=False)
def geocode(query: str):
    """
    Géocode robuste v19 :
    - Cas spécial : code postal FR seul (5 chiffres) -> CP, France
    - Détection automatique du pays pour les adresses étrangères
    - Répare les CP collés aux mots (ex : 'Hugo76600 le havre')
    """

    if not query or not isinstance(query, str):
        return None

    # Nettoyage de base
    q = clean_street_numbers(clean_internal_codes(_fix_postcode_spaces(_norm(query))))

    # Sépare les CP collés aux mots : "Hugo76600le" -> "Hugo 76600 le"
    q = re.sub(r"(\D)(\d{5})", r"\1 \2", q)
    q = re.sub(r"(\d{5})(\D)", r"\1 \2", q)

    q_low = q.lower().strip()

    # ================= 1) CAS SPECIAL : CP FR SEUL =================
    # Exemple : "33210" -> on force "33210, France"
    if re.fullmatch(r"\d{5}", q_low):
        geolocator = Nominatim(user_agent="moa_geo_cp_only")
        try:
            time.sleep(1)
            loc = geolocator.geocode(f"{q_low}, France", timeout=20, addressdetails=True)
        except Exception:
            loc = None

        if not loc:
            return None

        addr = loc.raw.get("address", {})
        country = addr.get("country", "France")
        postcode = addr.get("postcode", q_low)
        return (loc.latitude, loc.longitude, country, postcode)

    # ================= 2) DETECTION PAYS PAR HEURISTIQUES =================

    # 🇳🇱 Pays-Bas (CP 1234AB, villes NL, etc.)
    if re.search(r"\b\d{4}[a-z]{2}\b", q_low) or any(v in q_low for v in [
        "amsterdam", "rotterdam", "utrecht", "eindhoven", "groningen"
    ]):
        country_hint = "Netherlands"

    # 🇧🇪 Belgique (B3570, CP 4 chiffres, villes BE)
    elif (
        re.match(r"^b\d{4}$", q_low)
        or (re.fullmatch(r"\d{4}", q_low) and 1000 <= int(q_low) <= 9999)
        or any(v in q_low for v in ["belg", "aarschot", "alken", "ittre", "maasmechelen", "sambreville"])
    ):
        country_hint = "Belgium"

    # 🇱🇺 Luxembourg
    elif re.match(r"l-\d{4,5}", q_low) or "luxem" in q_low:
        country_hint = "Luxembourg"

    # 🇪🇸 Espagne (Vila-real, Castellón, 12540, Espagne…)
    elif (
        "vila-real" in q_low or "vilareal" in q_low
        or "castell" in q_low or "espa" in q_low
        or "barcelone" in q_low or "barcelona" in q_low
        or q_low.startswith("es-") or "12540" in q_low
    ):
        country_hint = "Spain"

    # 🇮🇹 Italie
    elif "ital" in q_low or q_low.startswith("it-") or any(v in q_low for v in [
        "brescia", "bedizzole", "milano", "roma", "verona"
    ]):
        country_hint = "Italy"

    # 🇨🇭 Suisse
    elif "suisse" in q_low or "switzerland" in q_low or "ch-" in q_low:
        country_hint = "Switzerland"

    # 🇫🇷 Défaut : France
    else:
        country_hint = "France"

    # ================= 3) REQUETE NOMINATIM =================

    # Si un pays est déjà écrit dans l’adresse, on ne rajoute rien
    query_full = q if has_explicit_country(q) else f"{q}, {country_hint}"

    geolocator = Nominatim(user_agent="moa_geo_v19")
    try:
        time.sleep(1)
        loc = geolocator.geocode(query_full, timeout=20, addressdetails=True)
        if not loc:
            return None
    except Exception:
        return None

    addr = loc.raw.get("address", {})
    country_res = addr.get("country", country_hint)
    cp_res = addr.get("postcode", "")

    # Ajustements fins

    # Vila-real -> toujours 12540 Espagne
    if "vila-real" in q_low or "vilareal" in q_low:
        cp_res = "12540"
        country_res = "Espagne"

    # Pays-Bas si CP 1234AB repéré
    if re.search(r"\b\d{4}[A-Za-z]{2}\b", q):
        country_res = "Pays-Bas"

    # Belgique si CP Bxxxx
    if re.match(r"^b\d{4}$", q_low):
        country_res = "Belgique"

    # Luxembourg si L-xxxx
    if re.match(r"l-\d{4}", q_low):
        country_res = "Luxembourg"

    return (loc.latitude, loc.longitude, country_res, cp_res)



def try_geocode_with_fallbacks(raw_addr: str, assumed_country_hint: str = "France"):
    """Essaye plusieurs variantes d'une même adresse pour fiabiliser le géocodage."""
    s = clean_street_numbers(clean_internal_codes(_fix_postcode_spaces(_norm(raw_addr))))
    explicit_overseas = has_explicit_country(s)

    g = geocode(s if explicit_overseas else f"{s}, {assumed_country_hint}")
    if g:
        return g

    cp, ville = extract_cp_city(s)
    if cp or ville:
        for variant in [f"{cp} {ville}", ville, cp]:
            g = geocode(variant + ("" if explicit_overseas else ", France"))
            if g:
                return g

    # Dernier essai brut
    return geocode(s)



 
def distance_km(base_coords, coords):
    """
    Calcule la distance entre deux points :
    1️⃣ Priorité : distance routière via OSRM (gratuite et sans clé)
    2️⃣ Fallback : distance géodésique (vol d’oiseau)
    Retourne un tuple : (distance_km arrondie, type_utilisé)
    """
    if not coords or not base_coords:
        return None, ""

    import requests
    from geopy.distance import geodesic

    try:
        # 🚗 Requête vers OSRM (service public)
        url = f"http://router.project-osrm.org/route/v1/driving/{base_coords[1]},{base_coords[0]};{coords[1]},{coords[0]}?overview=false"
        r = requests.get(url, timeout=15)
        if r.status_code == 200:
            js = r.json()
            d = js["routes"][0]["distance"] / 1000.0
            return round(d, 1), "API OSRM"
        else:
            print(f"⚠️ OSRM renvoie un code {r.status_code}")
    except Exception as e:
        print(f"⚠️ OSRM échouée : {e}")

    # 🕊️ Fallback vol d’oiseau
    d = geodesic(base_coords, coords).km
    return round(d, 1), "Vol d’oiseau"





# ================= COLONNES & CONTACT MOA (v12-style+) ======
def _find_columns(cols):
    """
    Détection robuste des colonnes :
    - champs clés (raison/catégorie/référent/email_referent/adresse)
    - groupes de colonnes contacts (tech/dir/comce/com)
    - colonnes 'contacts' génériques
    """
    res = {
        "tech_cols": [], "dir_cols": [], "comce_cols": [], "com_cols": [], "contact_cols": []
    }
    for c in cols:
        cl = c.lower().strip()

        # clés fixes
        if "raison" in cl and "sociale" in cl: res["raison"] = c
        elif "catég" in cl or "categorie" in cl: res["categorie"] = c
        elif ("référent" in cl and "moa" in cl) or ("referent" in cl and "moa" in cl): res["referent"] = c
        elif ("email" in cl and "referent" in cl) or ("email" in cl and "référent" in cl): res["email_referent"] = c
        elif "adress" in cl: res["adresse"] = c

        # contacts : large
        # on classe par priorité via mots-clés
        if "tech" in cl:
            res["tech_cols"].append(c)
        if "dir" in cl or "dirige" in cl:
            res["dir_cols"].append(c)
        if "comce" in cl:  # si tu as cet acronyme précis
            res["comce_cols"].append(c)
        # "com" peut être ambigu (company). On limite aux variantes usuelles:
        if re.search(r"\bcom\b|\bcommercial", cl):
            res["com_cols"].append(c)
        # colonnes génériques "contact" (si pas déjà rangées)
        if "contact" in cl and c not in (res["tech_cols"] + res["dir_cols"] + res["comce_cols"] + res["com_cols"]):
            res["contact_cols"].append(c)

        # colonne simple "contacts"
        if "contacts" == cl or cl.startswith("contacts "):
            res["contacts"] = c

    return res

def _first_email_in_text(text:str)->str|None:
    if not isinstance(text,str): return None
    m = EMAIL_RE.search(text)
    return m.group(0) if m else None

def _email_local(e:str)->str:
    return e.split("@",1)[0].lower() if isinstance(e,str) else ""

def _tokens(name:str)->list[str]:
    if not isinstance(name,str): return []
    return [t for t in re.split(r"[\s\-]+", name.lower()) if len(t)>=2]

def _emails_from_columns(row, cols):
    for col in cols:
        val = str(row.get(col, "")).strip()
        if not val: 
            continue
        em = _first_email_in_text(val) or (val if "@" in val else None)
        if em:
            return em
    return None

def choose_contact_moa(row, colmap):
    """
    Priorité:
      1) email_referent direct
      2) matching nom référent sur groupes Tech/Dir/Comce/Com (v12-style: colonnes nommées librement)
      3) fallback premier dispo Tech -> Dir -> Comce -> Com -> Contacts génériques (y compris "Contacts")
    """
    # 1) email référent explicite
    if colmap.get("email_referent"):
        v = row.get(colmap["email_referent"], "")
        if isinstance(v,str) and "@" in v:
            return v.strip()

    # groupes détectés
    tech = colmap.get("tech_cols", [])
    diro = colmap.get("dir_cols", [])
    comce = colmap.get("comce_cols", [])
    com = colmap.get("com_cols", [])
    generic = colmap.get("contact_cols", [])
    contacts_simple = [colmap.get("contacts")] if colmap.get("contacts") else []

    # 2) matching par nom du référent (si fourni)
    referent = str(row.get(colmap.get("referent",""), "")).strip() if colmap.get("referent") else ""
    toks = _tokens(referent)

    if toks:
        # on rassemble les candidats (ordre de priorité)
        scan_groups = [tech, diro, comce, com, generic, contacts_simple]
        for group in scan_groups:
            # on cherche l'email dont la partie locale match le plus de tokens
            best_email, best_score = None, -1
            for col in group:
                val = str(row.get(col, "")).strip()
                em = _first_email_in_text(val) or (val if "@" in val else None)
                if not em: 
                    continue
                local = _email_local(em)
                score = sum(t in local for t in toks)
                if score > best_score:
                    best_score, best_email = score, em
            if best_email and best_score > 0:
                return best_email

    # 3) fallback: premier email dispo selon l'ordre Tech -> Dir -> Comce -> Com -> Contacts génériques -> "Contacts"
    for group in [tech, diro, comce, com, generic, contacts_simple]:
        em = _emails_from_columns(row, group)
        if em:
            return em

    return ""
 
def process_csv_to_df(csv_bytes):
    """
    Lit le CSV et construit le DataFrame de base :
    - conserve les colonnes essentielles (raison, catégorie, adresse, référent)
    - calcule le Contact MOA selon la logique élargie (v12-style)
    - garde les colonnes d'implantations industrielles et du siège pour la sélection des sites
    - crée toujours une colonne 'Adresse' même si elle n’existe pas dans le CSV
    """
    try:
        df = pd.read_csv(csv_bytes, sep=None, engine="python")
    except Exception:
        df = pd.read_csv(csv_bytes, sep=";", engine="python")

    # Détection des colonnes importantes
    colmap = _find_columns(df.columns)

    out = pd.DataFrame()

    # --- Colonnes principales ---
    out["Raison sociale"] = (
        df[colmap.get("raison", "")].astype(str).fillna("")
        if colmap.get("raison") else df.get("Raison sociale", "")
    )

    out["Référent MOA"] = (
        df[colmap.get("referent", "")].astype(str).fillna("")
        if colmap.get("referent") else df.get("Référent MOA", "")
    )

    out["Catégories"] = (
        df[colmap.get("categorie", "")].astype(str).fillna("")
        if colmap.get("categorie") else df.get("Catégories", "")
    )

    # --- Adresse principale : crée toujours la colonne ---
    if colmap.get("adresse"):
        out["Adresse"] = df[colmap["adresse"]].astype(str).fillna("")
    elif "Adresse" in df.columns:
        out["Adresse"] = df["Adresse"].astype(str).fillna("")
    elif "Adresse-du-siège" in df.columns:
        out["Adresse"] = df["Adresse-du-siège"].astype(str).fillna("")
    elif "adresse-du-siège" in df.columns:
        out["Adresse"] = df["adresse-du-siège"].astype(str).fillna("")
    else:
        # dernier recours : première adresse industrielle trouvée
        possible_cols = [c for c in df.columns if "implant" in c.lower()]
        if possible_cols:
            out["Adresse"] = df[possible_cols[0]].astype(str).fillna("")
        else:
            out["Adresse"] = ""

    # --- Contact MOA (calcul automatique) ---
    out["Contact MOA"] = df.apply(lambda r: choose_contact_moa(r, colmap), axis=1)

    # --- Colonnes supplémentaires : implantations industrielles et siège ---
    extra_cols = []
    for c in df.columns:
        cl = str(c).lower()
        if ("implant" in cl and "indus" in cl) or ("siège" in cl) or ("siege" in cl):
            extra_cols.append(c)
    for c in extra_cols:
        out[c] = df[c].astype(str).fillna("")

    return out

def pick_site_with_indus_priority(addr_field: str, base_coords: tuple[float, float], row=None):
    """
    Priorité stricte :
      1) entreprises à adresse fixe (forçages)
      2) implantations industrielles
      3) siège
      4) fallback adresse principale
    Retour : (adresse, (lat,lon) or None, pays, cp, dist)
    """

    from geopy.distance import geodesic
    import re

    if row is None:
        return (addr_field or "").strip(), None, "", "", None

    name = str(row.get("Raison sociale", "") or "").lower().strip()

    # ---------------------------------------------------------------------
    # VALIDATION ADRESSES
    # ---------------------------------------------------------------------
    def _is_valid_address(a):
        if not isinstance(a, str):
            return False
        a = a.strip()
        if a in ["", "nan"]:
            return False
        if re.fullmatch(r"\d{5}\.0", a):
            return False
        if re.fullmatch(r"\d{5}", a):  # CP FR seul
            return False
        if re.fullmatch(r"\d{4}[A-Za-z]{2}", a):  # NL
            return False
        if re.fullmatch(r"[Bb]\d{4}", a):  # BE Bxxxx
            return False
        if re.fullmatch(r"[Ll]-\d{4,5}", a):  # LU
            return False
        if re.fullmatch(r"\d+", a):  # nombre seul
            return False
        return True

    # ---------------------------------------------------------------------
    # FIXED SITES
    # ---------------------------------------------------------------------
    FIXED_SITES = {
        "cci france pays-bas": ("16 Hogehilweg, 1101CD Amsterdam, Pays-Bas", "Pays-Bas", "1101CD"),
        "ecococon": ("Voderady 91942, Slovaquie", "Slovaquie", "91942"),
        "gramitherm": ("Boulevard de l’Europe 87, 5060 Sambreville, Belgique", "Belgique", "5060"),
        "litobox": ("Industriezone Kolmen, Stationsstraat 110bus2, B3570 Alken, Belgique", "Belgique", "B3570"),
        "takki": ("Rue du Halage 13, 1460 Ittre, Belgique", "Belgique", "1460"),
        "easy’go wood": ("Rue du Halage 13, 1460 Ittre, Belgique", "Belgique", "1460"),
        "easy'go wood": ("Rue du Halage 13, 1460 Ittre, Belgique", "Belgique", "1460"),
        "vandersanden": ("Slakweidestraat 41, 3630 Maasmechelen, Belgique", "Belgique", "3630"),
        "hekipia": ("69380 Chessy, Rhône, France", "France", "69380"),
        "eurocomponent": ("Via Malignani 10, 33058 San Giorgio di Nogaro, Italie", "Italie", "33058"),
        "eurocomposant": ("Via Malignani 10, 33058 San Giorgio di Nogaro, Italie", "Italie", "33058"),
        "retrofitt": ("Nieuwlandlaan 39/B224, 3200 Aarschot, Belgique", "Belgique", "3200"),
        "porcelanosa": ("Carretera Nacional 340, km 55,8, 12540 Vila-real, Espagne", "Espagne", "12540"),
        "butech": ("Carretera Nacional 340, km 55,8, 12540 Vila-real, Espagne", "Espagne", "12540"),
    }

    for k, (forced_addr, forced_country, forced_cp) in FIXED_SITES.items():
        if k in name:
            g = try_geocode_with_fallbacks(forced_addr, forced_country)
            if g:
                lat, lon, _, _ = g
                dist = geodesic(base_coords, (lat, lon)).km
                return forced_addr, (lat, lon), forced_country, forced_cp, dist
            return forced_addr, None, forced_country, forced_cp, None

    # ---------------------------------------------------------------------
    # NORMALISATION
    # ---------------------------------------------------------------------
    def _normalize(a):
        a = str(a or "")
        a = re.sub(r"multi[-\s]*sites?", "", a, flags=re.I)
        a = re.sub(r"\(.*?\)", "", a)
        a = re.sub(r"\s{2,}", " ", a).strip(" ,")
        if "chessy" in a.lower() and "69380" in a and "rhône" not in a.lower():
            a = "69380 Chessy, Rhône, France"
        return a

    # ---------------------------------------------------------------------
    # MULTI-SITE
    # ---------------------------------------------------------------------
    def _split_multisite(a):
        parts = re.split(r"[;\n/]", str(a or ""))
        return [p.strip(" ,") for p in parts if _is_valid_address(p.strip())]

    # ---------------------------------------------------------------------
    # AUTO-COERCION PAYS
    # ---------------------------------------------------------------------
    def _coerce_country(addr, country, cp):
        s = addr.lower()
        if cp.lower().startswith("b") and cp[1:].isdigit():
            return "Belgique"
        if re.fullmatch(r"\d{4}[a-z]{2}", cp.lower()):
            return "Pays-Bas"
        if cp.startswith("L-"):
            return "Luxembourg"
        if "vila-real" in s or cp == "12540":
            return "Espagne"
        if "ital" in s:
            return "Italie"
        return country or "France"

    # ---------------------------------------------------------------------
    # GEOCODE
    # ---------------------------------------------------------------------
    def _geocode_addr(a):
        g = try_geocode_with_fallbacks(a, "France")
        if not g:
            return None
        lat, lon, country, cp = g
        country = _coerce_country(a, country, cp)
        return (a, (lat, lon), country, cp)

    # ---------------------------------------------------------------------
    # BEST CANDIDATE
    # ---------------------------------------------------------------------
    def _best_of(lst):
        best = None
        for raw in lst:
            norm = _normalize(raw)
            g = _geocode_addr(norm)
            if not g:
                continue
            addr2, coords, country, cp = g
            dist = geodesic(base_coords, coords).km
            if country == "Espagne":
                cp = "12540"
            cand = (addr2, coords, country, cp, dist)
            if best is None or dist < best[-1]:
                best = cand
        return best

    # 1) IMPLANTATIONS
    indus_cols = [c for c in row.index if "implant" in c.lower() and "indus" in c.lower()]
    indus_list = []
    for c in indus_cols:
        indus_list += _split_multisite(row[c])
    best = _best_of(indus_list)
    if best:
        return best

    # 2) SIÈGE
    siege_cols = [c for c in row.index if "siège" in c.lower() or "siege" in c.lower()]
    siege_list = []
    for c in siege_cols:
        siege_list += _split_multisite(row[c])
    best = _best_of(siege_list)
    if best:
        return best

    # 3) ADRESSE PRINCIPALE
    norm = _normalize(addr_field)
    g = _geocode_addr(norm)
    if g:
        addr2, coords, country, cp = g
        dist = geodesic(base_coords, coords).km
        return addr2, coords, country, cp, dist

    return addr_field, None, "", "", None


# =================== DISTANCES & FINALE =====================
def compute_distances(df, base_address):
    """
    Adresse du projet : CP seul, CP+Ville, Ville ou adresse complète.
    Toujours géocodable via fallback solide.
    """

    if not base_address.strip():
        st.warning("⚠️ Aucune adresse de référence fournie.")
        return df, None, {}

    q = _fix_postcode_spaces(_norm(base_address))
    base = None

    # ======================================================
    # 1) CAS LE PLUS SIMPLE : CP seul → toujours accepté
    # ======================================================
    if re.fullmatch(r"\d{5}", q):
        base = geocode(f"{q}, France")
        if base:
            st.info(f"📍 Lieu interprété comme : {q}, France")

    # ======================================================
    # 2) CP + Ville OU Ville seule
    # ======================================================
    if not base:
        base = geocode(q)
        if base:
            st.info(f"📍 Lieu interprété comme : {q}")

    # ======================================================
    # 3) Fallback automatique CP/Ville
    # ======================================================
    if not base:
        cp, ville = extract_cp_city(q)

        if cp and ville:
            base = geocode(f"{cp} {ville}, France")
        elif cp:
            base = geocode(f"{cp}, France")
        elif ville:
            base = geocode(f"{ville}, France")

        if base:
            st.info(f"ℹ️ Lieu interprété comme fallback : {cp or ''} {ville or ''}".strip())

    # ======================================================
    # 4) ERREUR SI RIEN
    # ======================================================
    if not base:
        st.warning(f"⚠️ Lieu de référence non géocodable : '{base_address}'.")
        df2 = df.copy()
        df2["Pays"] = ""
        df2["Code postal"] = df2["Adresse"].apply(extract_cp_fallback)
        df2["Distance au projet"] = ""
        df2["Type de distance"] = ""
        df2["Fiabilité géocode"] = ""
        return df2, None, {}

    # ======================================================
    # 5) BASE OK → lancement distances
    # ======================================================
    base_coords = (base[0], base[1])
    chosen_coords = {}
    chosen_rows = []

    for _, row in df.iterrows():
        name = str(row.get("Raison sociale", "")).strip()
        adresse = str(row.get("Adresse", ""))

        kept_addr, coords, country, cp, best_dist = pick_site_with_indus_priority(
            adresse, base_coords, row
        )

        if coords:
            dist, dist_type = distance_km(base_coords, coords)
        else:
            dist = round(best_dist) if best_dist else None
            dist_type = ""

        if coords:
            chosen_coords[name] = (coords[0], coords[1], country)

        chosen_rows.append({
            "Raison sociale": name,
            "Pays": country,
            "Adresse": kept_addr,
            "Code postal": cp,
            "Distance au projet": dist,
            "Catégories": row.get("Catégories", ""),
            "Référent MOA": row.get("Référent MOA", ""),
            "Contact MOA": row.get("Contact MOA", ""),
            "Type de distance": dist_type,
            "Fiabilité géocode": "indus",
        })

    return pd.DataFrame(chosen_rows), base_coords, chosen_coords


# ========================= EXCEL ============================
def to_excel(df, template=TEMPLATE_PATH, start=START_ROW):
    """Excel complet : Adresse / CP séparés + Contact MOA e-mail."""
    wb = load_workbook(template)
    ws = wb.worksheets[0]
    max_cols = 8
    for r in range(start, ws.max_row+1):
        for c in range(1, max_cols+1):
            ws.cell(r, c, value=None)
    for i, (_, r) in enumerate(df.iterrows(), start=start):
        ws.cell(i,1, r.get("Raison sociale",""))
        ws.cell(i,2, r.get("Pays",""))
        ws.cell(i,3, r.get("Adresse",""))
        ws.cell(i,4, r.get("Code postal",""))
        ws.cell(i,5, r.get("Distance au projet",""))
        ws.cell(i,6, r.get("Catégories",""))
        ws.cell(i,7, r.get("Référent MOA",""))
        ws.cell(i,8, r.get("Contact MOA",""))   # e-mail dans Excel
        ws.cell(i,9, r.get("Type de distance",""))
    bio = BytesIO(); wb.save(bio); bio.seek(0); return bio

def to_simple(df, template="doc_base_contact_simple.xlsx", start=11):
    """
    Génère le fichier 'contact simple' dans le modèle :
    Colonnes :
      A = Raison sociale
      B = Référent MOA
      C = Contact MOA
      D = Catégories
    Les lignes commencent à start (=11).
    """

    # ouverture modèle
    wb = load_workbook(template)
    ws = wb.active

    # on efface d'anciennes valeurs
    for r in range(start, ws.max_row + 1):
        for c in range(1, 5):
            ws.cell(r, c).value = None

    # remplissage
    for i, (_, row) in enumerate(df.iterrows(), start=start):
        ws.cell(i, 1, row.get("Raison sociale", ""))
        ws.cell(i, 2, row.get("Référent MOA", ""))
        ws.cell(i, 3, row.get("Contact MOA", ""))
        ws.cell(i, 4, row.get("Catégories", ""))

    # export
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ===================== CARTE (Folium) =======================
def make_map(df, base_coords, coords_dict, base_address):
    fmap = folium.Map(location=[46.6, 2.5], zoom_start=5, tiles="CartoDB positron", control_scale=True)
    if base_coords:
        folium.Marker(base_coords, icon=folium.Icon(color="red", icon="star"),
                      popup=f"<b>Projet</b><br>{base_address}",
                      tooltip="Projet").add_to(fmap)
    for _, r in df.iterrows():
        name = r.get("Raison sociale","")
        c = coords_dict.get(name)
        if not c: continue
        lat, lon, country = c
        addr = r.get("Adresse","")
        cp = r.get("Code postal","")
        folium.Marker([lat,lon],
            icon=folium.Icon(color="blue", icon="industry", prefix="fa"),
            popup=f"<b>{name}</b><br>{addr}<br>{cp or ''} — {country}",
            tooltip=name).add_to(fmap)
        folium.map.Marker(
            [lat, lon],
            icon=DivIcon(icon_size=(180,36), icon_anchor=(0,0),
                         html=f'<div style="font-weight:600;color:#1f6feb;white-space:nowrap;'
                              f'text-shadow:0 0 3px #fff;">{name}</div>')
        ).add_to(fmap)
    return fmap

def map_to_html(fmap):
    s = fmap.get_root().render().encode("utf-8")
    bio = BytesIO(); bio.write(s); bio.seek(0); return bio

# ======================== INTERFACE =========================
st.markdown("""

<style>

/* ================================
      THEME CLAIR FORCÉ
================================ */
html, body, .stApp {
    background: #FFFFFF !important;
    color: #000000 !important;
}

/* Désactivation totale du mode sombre */
@media (prefers-color-scheme: dark) {
    html, body, .stApp {
        background: #FFFFFF !important;
        color: #000000 !important;
    }
}

/* ================================
      TITRES - TEXTES
================================ */
h1, h2, h3, h4, h5, h6 {
    color: #0B1D4F !important;
    font-family: "Inter", sans-serif !important;
    font-weight: 700 !important;
}

label, p, span, div, textarea, input {
    color: #000000 !important;
    font-family: "Inter", sans-serif !important;
}

/* ================================
      BOUTONS (STYLE MODERNE)
================================ */
.stButton>button,
.stDownloadButton>button {
    background: #0B1D4F !important;     /* Bleu foncé */
    color: #FFFFFF !important;          /* Texte blanc */
    border-radius: 8px !important;
    padding: 0.5rem 1.2rem !important;
    border: none !important;
    font-weight: 600 !important;
}

/* Hover */
.stButton>button:hover,
.stDownloadButton>button:hover {
    opacity: 0.85 !important;
    color: #FFFFFF !important;
}

/* Correction Streamlit : texte interne dans un <p> → forcer blanc */
.stButton button *,
.stDownloadButton button * {
    color: #FFFFFF !important;
}

.stButton button p,
.stDownloadButton button p {
    color: #FFFFFF !important;
}

/* ================================
      INPUTS / FILE UPLOAD
================================ */
.stTextInput>div>div>input,
.stFileUploader>div>div {
    background-color: #ffffff !important;
    color: #000000 !important;
}

/* ================================
      RADIOS HORIZONTALES
================================ */
.stRadio > div {
    flex-direction: row !important;
    gap: 20px !important;
}

/* ================================
      DATAFRAME
================================ */
[data-testid="stDataFrame"] {
    color: black !important;
}

</style>

""", unsafe_allow_html=True)

# ===============================================================
# APP
# ===============================================================



st.title("📍Sortie excel, Contacter JAROD en cas de problème")
st.image("Conseil-noir.jpg", width=220)


mode = st.radio("Choisir le mode :", ["🧾 Mode simple", "🚗 Mode enrichi (distances + carte)"], horizontal=True)
base_address = st.text_input("🏠 Adresse du projet (CP + ville ou adresse complète)",
                             placeholder="Ex : 33210 Langon  •  ou  17 Boulevard Allende 33210 Langon")

file = st.file_uploader("📄 Fichier CSV", type=["csv"])

name_full   = st.text_input("Nom du fichier Excel complet (sans extension)", "Sourcing_MOA")
name_simple = st.text_input("Nom du fichier contact simple (sans extension)", "MOA_contact_simple")
name_map    = st.text_input("Nom du fichier carte HTML (sans extension)", "Carte_MOA")

generate_map = False
if mode == "🚗 Mode enrichi (distances + carte)":
    generate_map = st.button("🗺️ Générer la carte maintenant")

if file and (mode == "🧾 Mode simple" or base_address):
    try:
        with st.spinner("⏳ Traitement en cours..."):
            base_df = process_csv_to_df(file)       # ✅ Contact MOA e-mail déjà calculé (v12-style+)
            if mode == "🚗 Mode enrichi (distances + carte)":
                df, base_coords, coords_dict = compute_distances(base_df, base_address)
            else:
                df, base_coords, coords_dict = base_df.copy(), None, {}

        st.success("✅ Traitement terminé")

        # contact simple
        x1 = to_simple(base_df, template="doc_base_contact_simple.xlsx", start=11)
        st.download_button("⬇️ Télécharger le contact simple",
                   data=x1, file_name=f"{name_simple}.xlsx",
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        if mode == "🚗 Mode enrichi (distances + carte)":
            # Excel complet
            x2 = to_excel(df)
            st.download_button("⬇️ Télécharger l'Excel complet",
                               data=x2, file_name=f"{name_full}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # Carte à la demande
            if generate_map and base_coords:
                fmap = make_map(df, base_coords, coords_dict, base_address)
                htmlb = map_to_html(fmap)
                st.download_button("📥 Télécharger la carte (HTML)",
                                   data=htmlb, file_name=f"{name_map}.html", mime="text/html")
                st_html(htmlb.getvalue().decode("utf-8"), height=520)
                st.caption("🧭 Distances calculées à vol d’oiseau (géodésiques).")

        st.subheader("📋 Aperçu des données")
        st.dataframe(df.head(12))

    except Exception as e:
        st.error(f"Erreur : {e}")




